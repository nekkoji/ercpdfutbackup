import os
import urllib.parse
import pandas as pd
from datetime import datetime
from PyQt5.QtWidgets import QMessageBox
from openpyxl import load_workbook
from openpyxl.styles import Font

from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.authentication_context import AuthenticationContext


ctx = None
ctx_auth = None
site_url = ""
base_link = ""


def authenticate_user(link, username, password, status_label):
    global ctx, site_url, ctx_auth
    username = username.strip()
    password = password.strip()

    if not username or not password:
        show_warning("Please enter both username and password.")
        return

    sample_link = link.strip()
    if not sample_link:
        show_warning("Please enter a SharePoint folder link.")
        return

    try:
        parsed_url = urllib.parse.urlparse(sample_link)
        path_parts = parsed_url.path.strip("/").split("/")
        if "sites" in path_parts:
            idx = path_parts.index("sites")
            site_url = f"https://{parsed_url.netloc}/{'/'.join(path_parts[idx:idx+2])}"
        else:
            site_url = f"https://{parsed_url.netloc}"

        status_label.setText("🔐 Authenticating...")
        ctx_auth = AuthenticationContext(site_url)
        if not ctx_auth.acquire_token_for_user(username, password):
            show_error("Invalid credentials.", "Authentication Failed")
            status_label.setText("❌ Authentication failed.")
            return

        ctx = ClientContext(site_url, ctx_auth)
        status_label.setText("✅ Authenticated. Ready to extract.")

    except Exception as e:
        show_error(f"Authentication error:\n{str(e)}")
        status_label.setText("❌ Authentication error.")


def logout_user(status_label, username_field, password_field):
    global ctx, site_url, ctx_auth, base_link
    ctx = None
    ctx_auth = None
    site_url = ""
    base_link = ""
    status_label.setText("🔓 Logged out. Please authenticate again.")
    username_field.clear()
    password_field.clear()


def extract_and_save(link, save_dir, status_label):
    global ctx, site_url, base_link
    if ctx is None:
        show_warning("Please authenticate before extracting.")
        return

    link = link.strip()
    save_dir = save_dir.strip()
    if not link or not save_dir:
        show_warning("Please enter SharePoint link and save location.")
        return

    try:
        parsed_url = urllib.parse.urlparse(link)
        query_params = urllib.parse.parse_qs(parsed_url.query)
        full_folder_path = ""

        if "id" in query_params:
            folder_url_encoded = query_params["id"][0]
            full_folder_path = urllib.parse.unquote(folder_url_encoded).lstrip("/")
        else:
            full_folder_path = urllib.parse.unquote(parsed_url.path).lstrip("/")

        path_parts = full_folder_path.split("/")
        if "sites" in path_parts:
            idx = path_parts.index("sites")
            folder_path = "/".join(path_parts[idx+2:])
        else:
            folder_path = full_folder_path

        encoded_folder_path = "/".join([urllib.parse.quote(part) for part in folder_path.split("/")])
        base_link = f"{site_url}/{encoded_folder_path}/"
        status_label.setText("📁 Accessing folder...")

        folder = ctx.web.get_folder_by_server_relative_url(folder_path)
        files = folder.files
        ctx.load(files)
        ctx.execute_query()

        data = []
        for file in files:
            name = file.properties["Name"]
            if name.endswith(".pdf"):
                full_url = f"{base_link}{name}"
                data.append([name, full_url])

        if not data:
            show_info("No PDF files found in the specified folder.", "No PDFs Found")
            status_label.setText("ℹ️ No PDFs found.")
            return

        df = pd.DataFrame(data, columns=["File Name", "Hyperlink"])
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(save_dir, f"sharepoint_file_links_{timestamp}.xlsx")
        df.to_excel(filename, index=False)

        # Format hyperlinks in Excel
        wb = load_workbook(filename)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            cell.font = Font(color="0000FF", underline="single", name="Georgia")
            cell.value = f'=HYPERLINK("{cell.value}", "{ws.cell(row=row, column=1).value}")'
        wb.save(filename)

        status_label.setText(f"✅ Saved: {filename}")
        show_info(f"Excel file created:\n{filename}", "Success")

    except Exception as e:
        show_error(f"Error occurred:\n{str(e)}")
        status_label.setText("❌ Error occurred.")


# Helper message dialogs (can also be placed in a shared utils/dialogs.py)
def show_error(message, title="Error"):
    QMessageBox.critical(None, title, message)

def show_warning(message, title="Warning"):
    QMessageBox.warning(None, title, message)

def show_info(message, title="Info"):
    QMessageBox.information(None, title, message)
